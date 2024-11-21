from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Provided data
# Update the products_data dictionary with the new entries
products_data = {
    "CE285ACB435A/436A/278A": {"Shoptech": 3000},
    "136A W1360A": {"Shoptech": 4000, "Toner Cap": 3800, "Jumia": 5390},
    "12A Q2612A": {"Shoptech": 3000, "Digital Store": 1899, "Jumia": 2800, "Kilimall": 2500},
    "151A CF151A": {"Toner Cap": 6000},
    "05A/80A CE505A/CF280A": {"Shoptech": 3000, "Cartridges": 1800, "Tech Store": 3800},
    "106A W1106A": {"Jumia": 5000, "Toner Cap": 3000, "Saruk": 4999},
    "107A W1107A": {"Toner Cap": 3000, "Shoptech": 4000, "Jumia": 3990, "Novatech": 3500},
    "17A CF217A": {"Shoptech": 3000, "Digital Store": 2599, "Novatech": 2000, "Jumia": 3800, "Toner Cap": 2000},
    "19A CF219A": {"Shoptech": 3900, "Jumia": 3400, "Toner Cap": 6500, "Toner Cartridge": 3900, "Kilimall": 4000, "Novatech": 2500},
    "26A CF226A": {"Shoptech": 4000, "Jumia": 2800, "Toner Cap": 2500, "Digital Store": 2799},
    "30A CF230A": {"Shoptech": 3000, "Jumia": 2800, "Kilimall": 3000},
    "32A CF232A (Drum)": {"Jumia": 2600, "Kilimall": 3500},
    "35A/36A/85A CB435A/CE285A": {"Shoptech": 3000},
    "59A CF259A": {"Shoptech": 7000, "Jumia": 7200, "Toner Cap": 4500, "Kilimall": 6500},
    "59A CF259A (Without Chip)": {"N/A": 0},
    "83A CF283A": {"Shoptech": 2900, "Kilimall": 2800},
    "117A W2070A Black": {"Shoptech": 4500, "Jumia": 6459, "Novatech": 4000},
    "117A W2071A Cyan": {"Shoptech": 4500, "Jumia": 6174},
    "117A W2072A Yellow": {"Novatech": 4000, "Jumia": 6174},
    "117A W2073A Magenta": {"Novatech": 4000, "Jumia": 6174},
    "126A/130A CE310A/CF350A (Black)": {"Digital Store": 1999},
    "126A/130A CE312A/CF352A (Yellow)": {"Digital Store": 1099},
    "126A/130A CE313A/CF353A (Magenta)": {"Digital Store": 1099},
    "201A CF400A Black": {"Shoptech": 4000},
    "201A CF401A Cyan": {"Jumia": 3990, "Kilimall": 3300},
    "201A CF402A Yellow": {"Shoptech": 5000, "Toner Cartridge": 5000, "Jumia": 3190},
    "201A CF403A Magenta": {"Shoptech": 5000, "Toner Cartridge": 5000, "Jumia": 3190},
    "203A CF540A Black": {"Jumia": 3800},
    "203A CF541A Cyan": {"Jumia": 3800, "Kilimall": 3200},
    "203A CF542A Yellow": {"Jumia": 3800, "Kilimall": 3200},
    "203A CF543A Magenta": {"Jumia": 2990, "Kilimall": 3200},
    "205A CF530A Black": {"Cartridges": 1600, "Jumia": 3000},
    "205A CF531A Cyan": {"Jumia": 3150, "Kilimall": 3200},
    "205A CF532A Yellow": {"Jumia": 3150, "Shoptech": 5000},
    "205A CF533A Magenta": {"Jumia": 2990},
    "410A CF410A Black": {"Jumia": 2300},
    "410A CF411A Cyan": {"Jumia": 3800},
    "410A CF412A Yellow": {"Jumia": 3800, "Shoptech": 5000},
    "410A CF413A Magenta": {"Jumia": 3800, "Shoptech": 5000},
    "415A W2030A Black": {"N/A": 0},
    "415A W2031A Cyan": {"Shoptech": 7000},
    "415A W2032A Yellow": {"Shoptech": 7000},
    "415A W2033A Magenta": {"N/A": 0},
    "CF540A(203A) Black": {"Jumia": 3800, "Kilimall": 3200},
    "CF541X(203A) Cyan": {"Jumia": 3800},
    "CF542A(203A) Yellow": {"Jumia": 3800, "Kilimall": 3200},
    "CF543A(203A) Magenta": {"Jumia": 3800},
    "CF530A(205A) Black": {"Jumia": 3000},
    "CF531A(205A) Cyan": {"Jumia": 3150, "Kilimall": 3200},
    "CF532A(205A) Yellow": {"Jumia": 3150},
    "CF533A(205A) Magenta": {"Jumia": 2990},
    "128A CE320A Black": {"Jumia": 3800, "Kilimall": 3500},
    "128A CE321A Cyan": {"Jumia": 3800, "Kilimall": 3500},
    "128A CE322A Yellow": {"Jumia": 3800, "Kilimall": 3500},
    "128A CE323A Magenta": {"Jumia": 3200, "Kilimall": 3500},
    "131A CF210A Black": {"Jumia": 5000, "Kilimall": 4000},
    "131A CF211A Cyan": {"Jumia": 3500, "Kilimall": 3500, "Shoptech": 4000},
    "131A CF212A Yellow": {"Jumia": 3500, "Kilimall": 4000, "Shoptech": 4000},
    "131A CF213A Magenta": {"N/A": 0},
    "207A W2210A Black": {"N/A": 0},
    "207A W2211A Cyan": {"N/A": 0},
    "207A W2212A Yellow": {"N/A": 0},
    "207A W2213A Magenta": {"N/A": 0},
    "C-CE314A/14K-DRUM": {"Shoptech": 4000},
    "C-CF219A/12K-DRUM": {"Jumia": 3400, "Novatech": 2500, "Kilimall": 4000, "Shoptech": 3900},
    "C-CF232A/23K-DRUM": {"Jumia": 3400, "Etech": 3500},
    "TK1120": {"Print Supply": 1500, "Etech": 1500, "Toners": 1840, "Fgee": 5000, "Shoptech": 1500, "MTech": 4500},
    "TK1140": {"Etech": 2000, "Toners": 3440, "Shoptech": 2200, "Innovative Computers": 1800, "Toner Cap": 2000},
    "TK1150": {"Etech": 2500, "Shoptech": 2400, "Toners": 3000, "Fgee": 6500},
    "TK1160": {"Innovative": 1800, "Shoptech": 3000, "Etech": 2500},
    "TK1170": {"Print Supply": 3000, "Etech": 2500, "Toners": 3340, "Shoptech": 2900},
    "TK4105": {"Etech": 3000, "Shoptech": 3000, "Innovative": 4000},
    "TK410": {
        "Etech": 3500,
        "Toners": 5040,
        "Centrifugal": 2500,
        "Innovative": 4000
    },
    "TK435": {
        "Shoptech": 3500,
        "MTech": 5300,
        "Etech": 3500
    },
    "TK475": {
        "Etech": 3500,
        "Fgee": 6000,
        "Toners": 5540
    },
    "TK675": {
        "Etech": 4000,
        "Digital Store": 4499,
        "Toners": 6940,
        "Shoptech": 3900,
        "Print Supply": 4500,
        "Centrifugal": 4500
    },
    "TK685": {
        "Shoptech": 3300,
        "MTech": 3000,
        "Innovative": 5700,
        "Cartridges": 3450
    },
    "TK715": {
        "Etech": 6000,
        "Shoptech": 8000,
        "Innovative": 6200,
        "Prodata": 7000,
        "Print Supply": 6000
    },
    "TK6115": {
        "Etech": 4000,
        "Shoptech": 3450,
        "Jumia": 3800
    },
    "TN 2305 Toner": {
        "Toners": 8500,
        "Print Supply": 10550,
        "Officemat": 7500,
        "Digital Store": 11500
    }
}

# Print the updated dictionary
print(products_data)


# Get all unique competitor names
all_competitors = set()
for competitors in products_data.values():
    all_competitors.update(competitors.keys())

# Create a new workbook
wb = Workbook()
ws = wb.active

# Write product numbers in Column A
ws.cell(row=1, column=1, value="Product Number")
ws.cell(row=1, column=2, value="Least Price")
ws.cell(row=1, column=3, value="Company with Least Price")

for idx, product_number in enumerate(products_data.keys(), start=2):
    ws.cell(row=idx, column=1, value=product_number)

# Write competitor names in Row 1, starting from Column D
for idx, competitor_name in enumerate(all_competitors, start=4):
    ws.cell(row=1, column=idx, value=competitor_name)

# Write competitor prices dynamically
for row_idx, (product_number, competitors) in enumerate(products_data.items(), start=2):
    lowest_price = min(competitors.values())
    product_with_lowest_price = [k for k, v in competitors.items() if v == lowest_price][0]
    
    ws.cell(row=row_idx, column=2, value=lowest_price)
    ws.cell(row=row_idx, column=3, value=product_with_lowest_price)
    
    for col_idx, competitor_name in enumerate(all_competitors, start=4):
        price = competitors.get(competitor_name, "")
        ws.cell(row=row_idx, column=col_idx, value=price)

# Add formula to calculate lowest price dynamically
for row_idx in range(2, len(products_data) + 2):
    ws[f"B{row_idx}"] = f"=MIN(D{row_idx}:{get_column_letter(len(all_competitors) + 3)}{row_idx})"

# Specify the path where you want to save the Excel file
file_path = "Advanced Trumark Asta_toners.xlsx"

# Save the workbook to the specified path
wb.save(file_path)

print(f"Excel file saved to: {file_path}")

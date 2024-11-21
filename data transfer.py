from openpyxl import load_workbook
from openpyxl import Workbook

def transfer_data(source_file, destination_file):
    # Load the source Excel file
    wb_source = load_workbook(filename=source_file)
    ws_source = wb_source.active  # Get the active worksheet

    # Create a new workbook for the destination file
    wb_dest = Workbook()
    ws_dest = wb_dest.active

    # Copy data from source to destination
    for row in ws_source.iter_rows(values_only=True):
        ws_dest.append(row)

    # Save the destination file
    wb_dest.save(destination_file)

    print("Data transferred successfully!")

# Paths for source and destination files
source_file = r'c:\Users\carso\Documents\June\simba general merchants limited.xlsx'  # Change this to your source Excel file
destination_file = r'c:\Users\carso\Documents\June\simba merchants products.xlsx'  # Change this to your destination Excel file

# Transfer data
transfer_data(source_file, destination_file)
